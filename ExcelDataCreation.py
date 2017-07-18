'''
Created on 18-Jul-2017

@author: BALASUBRAMANIAM
'''
from openpyxl import load_workbook
import random
filepath="G:/test/orangereport_2017.xlsx"
wb=load_workbook(filepath, read_only=False)
sheets=wb.get_sheet_names()
print(sheets)
sheet = wb.get_sheet_by_name('January_2017')
#write data in january 2017 work sheet

for row in range(1,10):
    for col in range(1,3):
        sheet.cell(column=col,row=row,value="%s" 
                   %str(random.randint(1,1000)) )
        
wb.save(filepath)
