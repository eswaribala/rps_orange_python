'''
Created on 17-Jul-2017

@author: BALASUBRAMANIAM
'''
from openpyxl import load_workbook

filePath="G:/test/annualReport_2017.xlsx"

wb= load_workbook(filePath, read_only=True)
sheetNames=wb.get_sheet_names()
#print(sheetNames)
sheet=wb.get_sheet_by_name('January_2017')
print(sheet['A4'].value)

for cell in range(1,sheet.max_row):

    print(cell,sheet.cell(row=cell,column=2).value)
    
    
    
